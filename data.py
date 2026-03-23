from openpyxl import Workbook, load_workbook
import os

excel_file = "vehicle_log.xlsx"

def save_to_excel(plate, status):

    # If file doesn't exist → create it
    if not os.path.exists(excel_file):

        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Log"

        # Header
        ws.append(["Plate Number", "Date", "Time", "Status"])

        wb.save(excel_file)

    # Load file
    wb = load_workbook(excel_file)
    ws = wb.active

    from datetime import datetime
    now = datetime.now()

    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # Add row
    ws.append([plate, date, time, status])

    wb.save(excel_file)